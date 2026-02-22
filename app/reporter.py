"""
Excel report generator.

Produces a multi-sheet .xlsx file:
    Sheet 1 – Summary by webmaster  (Module 1 / 2)
    Sheet 2 – 8-day scoring detail  (Module 3, one table per webmaster)
    Sheet 3 – Daily breakdown       (raw per-day stats per webmaster)
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app.scoring import ScoringResult, cohorts_to_dataframe


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_GREEN = "C6EFCE"
_YELLOW = "FFEB9C"
_RED = "FFC7CE"
_HEADER_FILL = "4472C4"
_HEADER_FONT_CLR = "FFFFFF"
_SUBHEADER_FILL = "D9E1F2"


def _score_fill(score_pct: float) -> str:
    if score_pct >= 90:
        return _GREEN
    if score_pct >= 70:
        return _YELLOW
    return _RED


def _write_header_row(ws, row_idx: int, values: list, bold: bool = True) -> None:
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.font = Font(bold=bold, color=_HEADER_FONT_CLR)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autofit(ws) -> None:
    """Set reasonable column widths."""
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws, summary_df: pd.DataFrame) -> None:
    ws.title = "Summary"
    headers = list(summary_df.columns)
    _write_header_row(ws, 1, headers)

    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    _autofit(ws)


def _write_scoring_sheet(ws, results: list[ScoringResult]) -> None:
    ws.title = "8-day Score"
    current_row = 1

    for result in results:
        # Webmaster header
        cell = ws.cell(row=current_row, column=1,
                       value=f"Webmaster: {result.webmaster}  |  "
                             f"Analysis date: {result.analysis_date}  |  "
                             f"Score: {result.score_pct:.1f}%")
        fill_color = _score_fill(result.score_pct)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(bold=True)
        current_row += 1

        cohort_df = cohorts_to_dataframe(result)
        if cohort_df.empty:
            ws.cell(row=current_row, column=1, value="No data in window")
            current_row += 2
            continue

        col_headers = list(cohort_df.columns)
        _write_header_row(ws, current_row, col_headers)
        current_row += 1

        for row in dataframe_to_rows(cohort_df, index=False, header=False):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=current_row, column=c_idx, value=val)
            current_row += 1

        # Totals row
        totals_cell = ws.cell(
            row=current_row, column=1,
            value=f"ИТОГО  |  Числитель: {result.numerator:.2f}  |  "
                  f"Знаменатель: {result.denominator:.2f}  |  Score: {result.score_pct:.2f}%"
        )
        totals_cell.font = Font(bold=True)
        totals_cell.fill = PatternFill("solid", fgColor=fill_color)
        current_row += 2  # blank line between webmasters

    _autofit(ws)


def _write_daily_sheet(ws, daily_df: pd.DataFrame, webmaster: str) -> None:
    ws.title = f"Daily_{webmaster[:20]}"
    title_cell = ws.cell(row=1, column=1, value=f"Daily breakdown — {webmaster}")
    title_cell.font = Font(bold=True)

    headers = list(daily_df.columns)
    _write_header_row(ws, 2, headers)

    for r_idx, row in enumerate(dataframe_to_rows(daily_df, index=False, header=False), start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    _autofit(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def save_report(
    output_path: str | Path,
    summary_df: pd.DataFrame,
    scoring_results: list[ScoringResult],
    daily_dfs: dict[str, pd.DataFrame],
) -> Path:
    """
    Write the full multi-sheet Excel report.

    Parameters
    ----------
    output_path      : destination .xlsx file path
    summary_df       : output of metrics.summary_by_webmaster()
    scoring_results  : list of ScoringResult from scoring.calc_score()
    daily_dfs        : {webmaster: daily_breakdown DataFrame}

    Returns the resolved output path.
    """
    output_path = Path(output_path)
    wb = Workbook()

    # Sheet 1 – Summary
    _write_summary_sheet(wb.active, summary_df)

    # Sheet 2 – 8-day scoring
    _write_scoring_sheet(wb.create_sheet(), scoring_results)

    # Sheet 3+ – Daily breakdown per webmaster
    for wm, ddf in daily_dfs.items():
        _write_daily_sheet(wb.create_sheet(), ddf, wm)

    wb.save(output_path)
    return output_path.resolve()
